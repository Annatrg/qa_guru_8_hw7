import os
from zipfile import ZipFile
import openpyxl
from PyPDF2 import PdfReader
import pandas


def test_xlsx():
    with ZipFile('tmp/files.zip') as file:
        archived_file = openpyxl.load_workbook(file.open('report_orders.xlsx')).active
        not_archived_file = openpyxl.load_workbook('resources/report_orders.xlsx').active
        assert archived_file.cell(row=2, column=11).value == not_archived_file.cell(row=2, column=11).value
        assert archived_file.max_column == not_archived_file.max_column
        assert archived_file.max_row == not_archived_file.max_row
        assert archived_file.max_row == not_archived_file.max_row


def test_xls():
    with ZipFile('tmp/files.zip') as file:
        archived_file = pandas.read_excel(file.open('delivered_not_payed.xls'))
        not_archived_file = pandas.read_excel('resources/delivered_not_payed.xls')
        assert archived_file.columns.any() == not_archived_file.columns.any()
        assert archived_file.shape[0] == not_archived_file.shape[0]
        assert file.getinfo('delivered_not_payed.xls').file_size == os.path.getsize('resources/delivered_not_payed.xls')


def test_txt():
    with ZipFile('tmp/files.zip') as file:
        assert file.read('test.txt').decode('utf-8') == 'test'
        assert file.getinfo('test.txt').file_size == os.path.getsize('resources/test.txt')


def test_pdf():
    with ZipFile('tmp/files.zip') as file:
        archived_pdf = PdfReader(file.open('tickets.pdf'))
        not_archived_pdf = PdfReader('resources/tickets.pdf')
        assert len(archived_pdf.pages) == len(not_archived_pdf.pages)
        assert file.getinfo('tickets.pdf').file_size == os.path.getsize('resources/tickets.pdf')
